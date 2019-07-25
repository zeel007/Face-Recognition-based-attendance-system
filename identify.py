import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook

try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string
import time

# get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename="reports.xlsx")
sheet = wb.get_sheet_by_name('16CE')


def getDateColumn():
    for i in range(1, len(sheet.rows[0]) + 1):
        col = get_column_letter(i)
        if sheet.cell('%s%s' % (col, '1')).value == currentDate:
            return col


Key = '1cb27881e27d42628ab2d61fcc5b14b2'
CF.Key.set(Key)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(60)]

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
    if filename.endswith(".jpg"):
        #imgurl = urllib.pathname2url(os.path.join(directory, filename))
        imgurl = "C:\Users\Mit Patel\Desktop\New folder\Autoattendance-Cognitive-master\Cropped_faces\\face1.jpg"
        res = CF.face.detect(imgurl)
        if len(res) != 1:
            print "No face detected."
            continue

        faceIds = []
        for face in res:
            faceIds.append(face['faceId'])
        res = CF.face.identify(faceIds, personGroupId)
        print filename
        print res
        for face in res:
            if not face['candidates']:
                print "Unknown"
            else:
                personId = face['candidates'][0]['personId']
                c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
                row = c.fetchone()
                attend[int(row[0])] += 1
                print row[1] + " recognized"
        time.sleep(6)

sheet = wb.active

columns = list(sheet)

for row in range(2,len(columns) + 1):
 #   rn = sheet.cell('A%s' % row).value
    rn = columns[row][0].value
    if rn is not None:
        rn = rn[-2:]
        if attend[int(rn)] != 0:
            col = getDateColumn()
            sheet['%s%s' % (col, str(row))] = 1

wb.save(filename="reports16CE.xlsx")
# currentDir = os.path.dirname(os.path.abspath(__file__))
# imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
# res = CF.face.detect(imgurl)
# faceIds = []
# for face in res:
#   faceIds.append(face['faceId'])

# res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
# print res
